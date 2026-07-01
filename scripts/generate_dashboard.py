#!/usr/bin/env python3
"""
generate_dashboard.py
Header-based parser — resolves columns by name so new columns never break it.
Falls back to positional parsing for legacy 26-col format.
"""
import json, os, io, re
from datetime import datetime
import openpyxl

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(REPO_ROOT, "data", "CIM_Master_Data.xlsx")
TEMPLATE  = os.path.join(REPO_ROOT, "template", "dashboard.html")
OUTPUT    = os.path.join(REPO_ROOT, "index.html")

REGION_NORMALIZE = {
    "WI / IL": "WI/IL",
    "Southeast": "Carolinas",
}
REGION_CONFIG = {
    "Colorado":   {"label": "Colorado",               "askNote": "regional ask"},
    "Minnesota":  {"label": "Minnesota",               "askNote": "regional ask"},
    "Florida":    {"label": "Florida",                 "askNote": "regional ask"},
    "WI/IL":      {"label": "Wisconsin / Illinois",    "askNote": "regional ask"},
    "Midwest":    {"label": "Midwest (IN / KY / OH)",  "askNote": "regional ask"},
    "Arizona":    {"label": "Arizona",                 "askNote": "regional ask"},
    "Texas":      {"label": "Texas",                   "askNote": "regional ask"},
    "Houston":    {"label": "Houston",                 "askNote": "regional ask"},
    "Carolinas":  {"label": "Carolinas (NC)",          "askNote": "ramping"},
    "California": {"label": "California",              "askNote": "pricing under review"},
}
REGION_ORDER = ["Colorado","Minnesota","Florida","WI/IL","Midwest",
                "Arizona","Texas","Houston","Carolinas","California"]


def normalize_region(raw, state):
    raw = str(raw or '').strip()
    if raw in ("MN / FL", "MN/FL"):
        return "Minnesota" if str(state).strip() == "MN" else "Florida"
    return REGION_NORMALIZE.get(raw, raw)

def sf(v):
    try: f = float(v); return None if f != f else f
    except: return None

def si(v):
    f = sf(v); return int(round(f)) if f is not None else None

def sy(v):
    if v is None: return None
    if isinstance(v, datetime): return str(v.year)
    m = re.search(r'\b(20\d{2})\b', str(v))
    return m.group(1) if m else (str(v).strip() or None)

def ss(v): return str(v).strip() if v else None

def rnd(v): return round(float(v), 4) if v is not None and sf(v) is not None else None

def mk_id(state, name):
    return f"{str(state).lower()}_{re.sub(r'[^a-z0-9]','',str(name).lower())}"

def fmt_lease(v):
    if v is None: return None
    if isinstance(v, datetime): return v.strftime("%b %Y")
    return str(v).strip()[:10] or None


# ── Header-based parser (v21+ with any column count) ─────────────────────────
# Maps logical field names to possible header strings (case-insensitive match)
HEADER_MAP = {
    "uid":       ["unit id"],
    "fullName":  ["name"],
    "shortName": ["location name"],
    "state":     ["state"],
    "region":    ["region / package", "region/package"],
    "type":      ["type"],
    "vintage":   ["vintage"],
    "suites":    ["suites"],
    "sqft":      ["sq ft", "sq. ft", "sqft"],
    "p3PaidOcc": ["p3 2026 paid occupancy"],
    "p3FilledOcc":["p3 2026 filled occupancy"],
    "p3Rev":     ["p3 2026 ltm rev ($k)", "p3 2026 ltm rev"],
    "p3EBITDA":  ["p3 2026 ltm ebitda ($k)", "p3 2026 ltm ebitda"],
    "p6PaidOcc": ["p6 2026 paid occ%", "p6 2026 paid occupancy"],
    "p6FilledOcc":["p6 2026 filled occ%", "p6 2026 filled occupancy"],
    "awr":       ["most recent weekly average rate", "awr"],
    "p12PaidOcc":["p12 2026 paid occupancy"],
    "p12Rev":    ["ltm p12 2026 rev ($k)", "ltm p12 2026 rev"],
    "p12EBITDA": ["ltm p12 2026 ebitda ($k)", "ltm p12 2026 ebitda"],
    "matureAWR": ["mature weekly average rate", "mature awr"],
    "matureOcc": ["mature paid occupancy"],
    "matureRev": ["mature rev ($k)", "mature rev"],
    "matureEB":  ["mature ebitda ($k)", "mature ebitda"],
    "valP12":    ["valuation p12 2026"],
    "valMature": ["valuation mature"],
    "askPrice":  ["ask price ($k)", "ask price"],
    "openDate":  ["initial open date"],
    "leaseExp":  ["current lease expiration"],
    "options":   ["options"],
    "nextOpt":   ["next option expiration"],
    "lastOpt":   ["last option expiration"],
    "salons3mi": ["3 mile salon counts", "3 mile salons"],
    "pop3mi":    ["3 mile population"],
    "notes":     ["notes"],
    "expNotes":  ["expansion notes"],
}

def build_col_map(ws):
    """Return dict of field_name -> column_index (1-based) from header row."""
    raw = {ws.cell(2, c).value: c for c in range(1, ws.max_column + 1)
           if ws.cell(2, c).value}
    normalized = {str(k).lower().strip(): v for k, v in raw.items()}
    col = {}
    for field, candidates in HEADER_MAP.items():
        for candidate in candidates:
            if candidate in normalized:
                col[field] = normalized[candidate]
                break
    return col

def parse_header_based(ws):
    col = build_col_map(ws)
    print(f"  Mapped {len(col)}/{len(HEADER_MAP)} fields from headers")
    missing = [k for k in HEADER_MAP if k not in col]
    if missing: print(f"  Missing: {missing}")

    def g(r, field, default=None):
        c = col.get(field)
        return ws.cell(r, c).value if c else default

    locs = []
    for r in range(3, ws.max_row + 1):
        uid   = g(r, "uid")
        state = ss(g(r, "state"))
        name  = ss(g(r, "shortName")) or ss(g(r, "fullName"))
        if not uid or not name or not state: continue

        region = normalize_region(g(r, "region"), state)
        if region not in REGION_CONFIG: continue

        p3Rev  = sf(g(r, "p3Rev"));   p3EB  = sf(g(r, "p3EBITDA"))
        p12Rev = sf(g(r, "p12Rev"));  p12EB = sf(g(r, "p12EBITDA"))
        mRev   = sf(g(r, "matureRev")); mEB = sf(g(r, "matureEB"))
        ask    = sf(g(r, "askPrice"))

        # Skip locations with no financial data at all
        if ask is None and p3Rev is None and p12Rev is None:
            continue

        p6occ  = sf(g(r, "p6PaidOcc"))
        p3occ  = sf(g(r, "p3PaidOcc"))
        p12occ = sf(g(r, "p12PaidOcc"))
        mawocc = sf(g(r, "matureOcc"))

        notes = ss(g(r, "notes")) or ''
        expn  = ss(g(r, "expNotes")) or ''

        locs.append({
            "id": mk_id(state, name), "uid": uid,
            "name": name, "fullName": ss(g(r, "fullName")) or f"{state} | {name}",
            "state": state, "region": region,
            "vintage": si(g(r, "vintage")), "locType": ss(g(r, "type")) or '',
            "suites": si(g(r, "suites")), "sqft": si(g(r, "sqft")),
            "p6Occ":  p6occ,
            "p3Occ":  p3occ,  "p3Rev":  rnd(p3Rev),  "p3EBITDA":  rnd(p3EB),
            "awr":    rnd(sf(g(r, "awr"))),
            "p12Occ": p12occ, "p12Rev": rnd(p12Rev), "p12EBITDA": rnd(p12EB),
            "matureOcc": mawocc, "matureRev": rnd(mRev),
            "matureEB":  rnd(mEB), "matureAWR": rnd(sf(g(r, "matureAWR"))),
            "askPrice":  rnd(ask),
            "valP12":    rnd(sf(g(r, "valP12"))), "valMature": rnd(sf(g(r, "valMature"))),
            "leaseExp":  fmt_lease(g(r, "leaseExp")), "options": ss(g(r, "options")),
            "nextOpt":   sy(g(r, "nextOpt")),  "lastOpt": sy(g(r, "lastOpt")),
            "salons3mi": si(g(r, "salons3mi")), "pop3mi": si(g(r, "pop3mi")),
            "notes": notes[:80], "fullNotes": notes, "expansionNotes": expn,
            # Legacy aliases
            "occ": p6occ, "ltmRev": rnd(p3Rev), "ltmEBITDA": rnd(p3EB), "pfEBITDA": rnd(mEB),
        })
    return locs


# ── Legacy 26-column positional parser ───────────────────────────────────────
def parse_26(ws):
    locs = []
    for r in range(3, ws.max_row + 1):
        def g(c): return ws.cell(r, c).value
        uid = g(1); name = g(2); state = g(4)
        if not uid or not name or not state: continue
        region = normalize_region(g(5), state)
        if region not in REGION_CONFIG: continue
        state = str(state).strip()
        p3Rev = sf(g(13)); p3EB = sf(g(14)); mEB = sf(g(16))
        ask = sf(g(19)); notes = ss(g(25)) or ''
        locs.append({
            "id": mk_id(state, name), "uid": uid,
            "name": ss(name), "fullName": f"{state} | {ss(name)}",
            "state": state, "region": region,
            "vintage": si(g(8)), "locType": ss(g(7)) or '',
            "suites": si(g(9)), "sqft": si(g(10)),
            "p6Occ": sf(g(11)), "p3Occ": sf(g(11)), "p3Rev": rnd(p3Rev), "p3EBITDA": rnd(p3EB),
            "awr": None, "p12Occ": None, "p12Rev": None, "p12EBITDA": None,
            "matureOcc": None, "matureRev": None, "matureEB": rnd(mEB), "matureAWR": None,
            "askPrice": rnd(ask), "valP12": None, "valMature": rnd(sf(g(18))),
            "leaseExp": sy(g(22)), "options": None, "nextOpt": None, "lastOpt": None,
            "salons3mi": None, "pop3mi": None,
            "notes": notes[:80], "fullNotes": notes, "expansionNotes": "",
            "occ": sf(g(11)), "ltmRev": rnd(p3Rev), "ltmEBITDA": rnd(p3EB), "pfEBITDA": rnd(mEB),
        })
    return locs


def load_and_parse():
    with open(XLSX_PATH, 'rb') as f: raw = f.read()
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb['Locations'] if 'Locations' in wb.sheetnames else wb.active
    mc = ws.max_column
    if mc >= 35:
        print(f"Detected: header-based format  (max_col={mc})")
        return parse_header_based(ws)
    else:
        print(f"Detected: 26-column legacy  (max_col={mc})")
        return parse_26(ws)


def build_groups(locs):
    totals = {k: 0.0 for k in REGION_CONFIG}
    for l in locs:
        if l['region'] in totals and l['askPrice']:
            totals[l['region']] += l['askPrice']
    groups = []
    for key in REGION_ORDER:
        cfg = REGION_CONFIG[key]
        if not any(l['region'] == key for l in locs): continue
        t = totals[key]
        ask_str = f"${t/1000:.2f}M" if t > 0 else "TBD"
        groups.append({"key": key, "label": cfg["label"], "ask": ask_str,
                       "askNote": cfg["askNote"], "regions": [key]})
    return groups


def main():
    locs   = load_and_parse()
    groups = build_groups(locs)

    real       = [l for l in locs if l.get('id') != '_']
    with_data  = [l for l in real if l['p3Rev'] is not None]
    total_rev  = sum(l['p3Rev']    or 0 for l in with_data)
    total_eb   = sum(l['p3EBITDA'] or 0 for l in with_data)
    total_ask  = sum(l['askPrice'] or 0 for l in real if l['askPrice'])
    has_tbd    = any(not l['askPrice'] for l in real)

    stat_suites = f"{sum(l['suites'] or 0 for l in real):,}"
    stat_rev    = f"${total_rev/1000:.1f}M"
    stat_eb     = f"${total_eb/1000:.1f}M"
    stat_ask    = f"${total_ask/1000:.1f}M" + ("+" if has_tbd else "")

    data_block = (f"const locations = {json.dumps(locs, indent=2)};\n\n"
                  f"const REGION_GROUPS = {json.dumps(groups, indent=2)};\n")

    with open(TEMPLATE, 'r', encoding='utf-8') as f: tmpl = f.read()
    out = tmpl.replace('{{DATA_BLOCK}}', data_block)
    for pat, val in [
        (r'id="statLocs">[^<]*',   f'id="statLocs">{len(real)}'),
        (r'id="statSuites">[^<]*', f'id="statSuites">{stat_suites}'),
        (r'id="statRev">[^<]*',    f'id="statRev">{stat_rev}'),
        (r'id="statEBITDA">[^<]*', f'id="statEBITDA">{stat_eb}'),
        (r'id="statAsk">[^<]*',    f'id="statAsk">{stat_ask}'),
    ]:
        out = re.sub(pat, val, out)

    with open(OUTPUT, 'w', encoding='utf-8') as f: f.write(out)

    rc = {}
    for l in real: rc[l['region']] = rc.get(l['region'], 0) + 1
    print(f"index.html written  ({len(real)} locations, {len(groups)} region groups)")
    for key in REGION_ORDER:
        if key in rc: print(f"  {key}: {rc[key]} locations")

if __name__ == '__main__':
    main()
